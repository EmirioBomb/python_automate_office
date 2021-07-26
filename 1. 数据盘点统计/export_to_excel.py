# -*- coding: utf-8
#
# 数据治理 - 数据定标
# 主要根据数据盘点表中的基本数据，将数据填充到定标表对应的列中
# author: Emirio
# date: 2021.07.14
#
# 功能特点: 
#   1. 引用简单的UI模块，便于用户操作
#   2. 通过pandas强大的功能，简化Excel操作
#   3. 使用第三方库减少开发量
#
# 环境及库的安装:
#    1. 安装python3.8+(本项目基于3.8版本开发, 不保证低于此版本的python出现的BUG)
#    2. 安装easygui, 主要用来显示图形对话框: pip3 install easygui
#    3. 安装pandas, 主要处理Excel表格: pip3 install pandas
#    4. 安装pinyin, 主要实现数据项中文名称对应的数据项代码以拼单首字母展示(大写)，如: 编号 - BH


import os
import re
from sys import exit

import easygui as eg
import pandas as pd
import pinyin as py
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

import config as cf


def open_file():
    """
    打开文件，其中文件类型只限定于 *.xls、 * .xlsx、 *.csv等标准的Excel文件，且只支持选择一个文件

    :return String file_name: 选中的文件
    """
    file_name = eg.fileopenbox(title="Select Excel file like .xls/.xlsx to open",
                               default=".",
                               filetypes=["*.xlsx", "*.xls", "*.csv"],
                               multiple=False)
    if file_name is None:
        eg.msgbox("未选中任何文件，将结束本程序！", Warning)
        exit(0)
    return file_name


def read_excel_info(file_name):
    """
    读取表格文件，将数据按照指定的列进行分组

    :param String file_name: 文件名
    :return DataFrame: 分组后的表格数据
    """
    with open(file_name) as f:
        # 读取表格文件
        df = pd.read_excel(file_name).filter(items=cf.FILTERED_TABLE_HEADER)
        # csv格式处理效率会高一些？？？
        # df = pd.read_csv(file_name).filter(items=cf.FILTERED_TABLE_HEADER)

        # 将dataframe按照指定列进行分组
        grouped_df = df.groupby(cf.AREA_FLAG)
        return grouped_df


def get_initial(s, delimiter=''):
    """
    重写第三方库pinyin方法
    @overwrite get_initial(u(s), delimiter=' ')

    获取中文首字母对应的大写形式
    注意下述中使用: _pinyin_generator方法时要使用: py.pinyin._pinyin_generator()调用方式，而不能直接使用: py._pinyin_generator()

    :param String s: 中文字符串
    :param String delimiter: 分割符, 将原默认改为''，代替原来的' '
    :return String: 返回中文首字母的大写形式
    """
    # 需将先转成字符串格式，防止数值等一些类型的值为空数据(NAN)时导致程序异常
    if str(s) != "nan":  
        initials = (p[0] for p in py.pinyin._pinyin_generator(s, format="strip"))
        return delimiter.join(initials).upper()
    else:
        return ""


def get_py_name(item):
    """
    通过正则表达式过滤掉首字母中的特殊字符，只保留数字与英文字母

    :param String item: 中文字符串
    :return String: 格式化后的拼音首字母
    """
    # 获取默认中文名称对应的首字母的大写形式
    py_source_name = get_initial(item)

    # 过滤掉特殊字符，只保留数字与英文字母
    py_formatted_name = re.sub(r'[^a-zA-Z0-9]','', py_source_name)
    return py_formatted_name


def get_data_source(grouped_df):
    """
    获取可信数据源及数据分布，所谓可信息数据源只是在数据分布中筛选出[指定值]的单据类型

    :param Dataframe grouped_df: 分组后的数据
    :return List usage_list: 可信数据源或数据分布
    """
    source_list = grouped_df[cf.BILL_TYPE_FLAG].values
    usage_list = ",\n".join(source_list)
    return usage_list


def get_data(grouped_df):
    """
    将按照区域分组后的数据进行分类统计

    :param DataFrame grouped_df: 按区域分组后的数据
    :return Dict DATA_TABLE: 初始化DataFrame所需的数据表
    """
    for area_name, area_group in grouped_df:
        # 不统计业务隐藏字段区域
        if area_name == cf.HIDDEN_VALUE:
            continue
        # 获取: 数据项字段名，过滤掉指定列中重复的数据，并将结果转为list类型
        item_list = area_group[cf.FIELD_NAME_FLAG].drop_duplicates().tolist()
        # 如果数据项字段包括序号，则将其删除
        if cf.SERIAL_NO in item_list:
            item_list.remove(cf.SERIAL_NO)

        # 遍历数据项字段名
        for item in item_list:
            cf.EDIT_DATE_LIST.append("")                                    # 盘点日期
            cf.CODE_NUMBER_LIST.append("")                                  # 资产编号
            cf.ITEM_NAME_LIST.append(item)                                  # 数据项中文名称
            cf.ITEM_CODE_LIST.append(get_py_name(item))                     # 数据项代码，拼音首字母
            cf.ITEM_MEANING_LIST.append("")                                 # 数据项业务含义
            cf.AREA_NAME_LIST.append(area_name)                             # 区域
            cf.DEPT_NAME_LIST.append(cf.DEPT_NAME)                          # 归口管理部门
            cf.DATA_MANAGER_LIST.append(cf.DATA_MANAGER)                    # 数据管家
            cf.MAINTAINANCE_METHOD_LIST.append(cf.MAINTAINANCE_METHOD)      # 维护方式
            cf.DATA_TYPE_LIST.append(cf.DATA_TYPE)                          # 数据类型
            cf.DATA_FORMAT_LIST.append(cf.DATA_FORMAT)                      # 数据格式
            cf.IS_INNERBOUND_LIST.append(cf.IS_INNERBOUND)                  # 是否入仓
            cf.THEME_AREA_LIST.append(cf.THEME_AREA)                        # 主题域

            # 通过数据项字段名时进行分组，用于以下几列的数据获取
            source_group = area_group[area_group[cf.FIELD_NAME_FLAG] == item]

            cf.DISTRIBUTE_SOURCE_LIST.append(get_data_source(source_group)) # 数据分布
            # 通过字段名与使用对数据进行分组，获取可信数据源
            create_group = area_group[(area_group[cf.FIELD_NAME_FLAG] == item) & (area_group[cf.USAGE_FLAG] == cf.CREATE_VALUE)]
            cf.CREATE_SOURCE_LIST.append(get_data_source(create_group))     # 可信数据源

            value_range = "" if (len(source_group[cf.VALUE_RANGE].values) == 0) else source_group[cf.VALUE_RANGE].values[0]
            cf.VALUE_RANGE_LIST.append(value_range)                         # 取值范围， 默认设置为空

            default_value = "" if (len(source_group[cf.DEFAULT_VALUE].values) == 0) else source_group[cf.DEFAULT_VALUE].values[0]
            cf.DEFAULT_VALUE_LIST.append(default_value)                     # 字段默认值，默认设置为空

            field_contraint = "" if (len(source_group[cf.CONTRAINT_VALUE].values) == 0) else source_group[cf.CONTRAINT_VALUE].values[0]
            cf.FIELD_CONSTRAINT_LIST.append(field_contraint)                # 字段约束，默认设置为空

            security_level = cf.LEVEL_TWO if ((len(source_group[cf.SECURITY_LEVEL].values) == 0) or (str(source_group[cf.SECURITY_LEVEL].values[0]) == "nan")) else source_group[cf.SECURITY_LEVEL].values[0]
            cf.SECURITY_LEVEL_LIST.append(security_level)                   # 安全级别，默认设置为L2

            reference_standard = cf.MADE_BY_YQJR if (len(source_group[cf.REFERENCE_STANDARD].values) == 0) else source_group[cf.REFERENCE_STANDARD].values[0]
            cf.REFERENCE_STANDARD_LIST.append(reference_standard)           # 参考标准，默认设置为我司实践
    return cf.DATA_TABLE


def check_file():
    """
    检查当前目录下将要导出的文件是否已存在，若存在提示用户是否覆盖文件，若不存在则创建一个空文件
    """
    if os.path.isfile(cf.ABS_FILE_NAME):
        message = "当前目录下已有[" + cf.FILE_NAME + "]文件，是否要覆盖原文件？"
        overwrite = eg.ccbox(msg=message, title="提示")
        if not overwrite:
           eg.msgbox(msg="结束本次操作，请修改当前目录下的文件名后再重试！")
    else:
        with open(cf.ABS_FILE_NAME, mode="w"): pass


def initiate_cell_style(writer, file_name):
    """
    单元格样式格式化，主要添加边框，字体颜色，对齐方式，自动调整列宽度等.
    其中自动调整列宽度只针对于常规默认字体大小(10)，字体大小若超过默认值，则需要重新设计自动调整列宽度算法
    
    :param ExcelWriter writer: 表格写入对象
    :param String file_name: 要保存的文件名
    """
    # 获取工作簿对象
    workbook = writer.book
    # 启用当前工作表，可用 workbook.worksheets[0]代替active
    worksheet = workbook.active

    # 初始化单元络边框粗细与颜色
    side = Side(style='thin', color='000000')
    # 初始化单元格边框样式
    border = Border(top=side, bottom=side, left=side, right=side)
    # 初始化字体格式
    font = Font(color='FF0000', bold=True, size=10)
    # 初始化单元格对齐方式
    alignment = Alignment(horizontal="center", vertical="center")

    counter = 0                 # 计数器，第一行为主题区，单独设置样式
    # 遍历单元格并初始化其样式
    for row in worksheet:
        counter += 1
        for cell in row:
            # 如果单元格数据为空时，不添加边框
            # if worksheet[cell.coordinate].value:
            #     worksheet[cell.coornidate].border = border
            if counter == 1:    # 第一行为主题区，格式化字体大小与颜色
                worksheet[cell.coordinate].font = font
            worksheet[cell.coordinate].border = border
            worksheet[cell.coordinate].alignment = alignment

    # 自动调整每列的宽度，下述方法仅适用于字段大小为常规大小(10)，字段过大无法自动适应
    # 临时解决方案: 通过计算当前字体大小后再进行适当的扩大最大长度，达到自动调整宽度的目的
    for i in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(i)].bestFit = True
        worksheet.column_dimensions[get_column_letter(i)].auto_size = True
    # 保存当前工作簿
    workbook.save(file_name)


def export_excel(df, file_name):
    """
    导出excel文件
    """
    with pd.ExcelWriter(file_name, mode="w", engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=cf.SHEET_NAME, index=False)
            initiate_cell_style(writer, cf.ABS_FILE_NAME)


def main():
    # 1. 打开文件
    file_name = open_file()
    # 2. 读取表格数据
    grouped_df = read_excel_info(file_name)
    # 3. 交数据分类统计后，初始化为pandas所需的数据框
    df = pd.DataFrame(get_data(grouped_df))
    # 4. 检查要导出的文件是否已存在
    check_file()
    # 5. 导出Excel文件
    export_excel(df, cf.ABS_FILE_NAME)
    # 6. 提示信息
    eg.msgbox("成功导出数据，请查看文件: " + cf.FILE_NAME)


if __name__ == '__main__':
    try:
        main()
    except:
        eg.exceptionbox()
